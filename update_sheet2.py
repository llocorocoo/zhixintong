import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# File paths
PRIMARY_PATH = r"E:\claude\zhixintong\评分细则.xlsx"
FALLBACK_PATH = r"E:\claude\zhixintong\评分细则_tmp.xlsx"

# Colors
HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"

# Section fill colors (override zebra)
SECTION_FILLS = {
    "一": "FFF2CC",   # 真实性核查 - yellow-tint
    "二": "E2EFDA",   # 稳定性评估 - green-tint
    "三": "FCE4D6",   # 合规性核查 - orange-tint
    "四": "DDEBF7",   # 安全性核查 - blue-tint
    "五": "EAD1DC",   # 专业性认证 - pink-tint
}

# Data rows
DATA = [
    ("一、真实性核查", "身份核实", "始终显示",
     "姓名\n性别\n年龄\n证件号码\n初始发证地",
     "固定展示"),
    ("一、真实性核查", "学历核实", "始终显示",
     "学校名称\n专业\n学历层次",
     "每条记录一行展示"),
    ("一、真实性核查", "工作经历核实", "始终显示",
     "公司名称\n职位\n起止时间",
     "每条记录一行展示"),
    ("二、稳定性评估", "工作稳定性", "始终显示",
     "平均在职时长\n过去5年跳槽次数\n最长一段工作年限",
     "固定展示"),
    ("三、合规性核查", "竞业限制", "有协议记录（无被诉）",
     "涉及企业\n协议状态",
     "命中必显"),
    ("三、合规性核查", "竞业限制", "有协议记录且存在违约被诉",
     "涉及企业\n协议状态\n案号\n案由\n当事人身份\n审理法院",
     "命中必显；被诉时加诉讼字段"),
    ("三、合规性核查", "前雇主违纪记录", "命中",
     "涉及企业\n违纪情况",
     "命中必显"),
    ("三、合规性核查", "行政合规记录", "已修复记录",
     "处罚机关\n处罚内容",
     "命中必显；修复记录不显示日期"),
    ("三、合规性核查", "行政合规记录", "公示期内记录",
     "处罚机关\n处罚内容\n处罚日期\n负面类型",
     "命中必显"),
    ("四、安全性核查", "司法诉讼", "命中",
     "案件编号\n案由\n当事人身份\n审理法院\n结案方式",
     "命中必显"),
    ("四、安全性核查", "法院被执行人", "已结案",
     "案件编号\n执行法院\n案件状态",
     "命中必显"),
    ("四、安全性核查", "法院被执行人", "执行中",
     "案件编号\n执行法院\n执行标的\n案件状态\n立案时间",
     "命中必显；执行中加标的金额与立案时间"),
    ("四、安全性核查", "犯罪记录", "未发现",
     "固定文案：未发现犯罪记录",
     "始终显示结论"),
    ("四、安全性核查", "关联企业", "存在异常企业",
     "企业名称\n担任职务\n企业状态",
     "命中必显"),
    ("四、安全性核查", "劳动争议记录", "命中",
     "争议记录\n核实说明",
     "命中必显"),
    ("四、安全性核查", "行业违规记录", "命中",
     "风险等级\n处罚时间\n处罚摘要",
     "命中必显"),
    ("五、专业性认证", "职业资格证书", "始终显示",
     "证书名称（颁发机构）",
     "固定展示"),
    ("五、专业性认证", "学历层次", "始终显示",
     "学历层次",
     "固定展示"),
]

# Border style
thin_side = Side(style='thin', color='000000')
thin_border = Border(
    left=thin_side, right=thin_side,
    top=thin_side, bottom=thin_side
)

def get_section_key(dim_text):
    """Extract section number (一/二/三/四/五) from 维度 text."""
    for key in SECTION_FILLS:
        if dim_text.startswith(key):
            return key
    return None

def build_sheet(ws):
    # Clear all content
    ws.delete_rows(1, ws.max_row + 1)

    # Remove any existing freeze
    ws.freeze_panes = None

    # Set column widths
    col_widths = [18, 18, 28, 60, 35]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Header row ---
    headers = ["维度（大类）", "调查项", "子情景/触发条件", "报告展示字段", "备注"]
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color=HEADER_FG, name="微软雅黑", size=11)
    header_align = Alignment(horizontal="center", vertical="center",
                              wrap_text=True)

    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 22

    # --- Data rows ---
    for row_idx, row_data in enumerate(DATA, start=2):
        dim, item, condition, fields, remark = row_data
        values = [dim, item, condition, fields, remark]

        # Determine fill color based on section
        section_key = get_section_key(dim)
        fill_color = SECTION_FILLS.get(section_key, "FFFFFF")
        row_fill = PatternFill(fill_type="solid", fgColor=fill_color)

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(
                horizontal="left", vertical="center",
                wrap_text=True
            )
            cell.border = thin_border

        ws.row_dimensions[row_idx].height = 40

    # Freeze pane at A2
    ws.freeze_panes = "A2"


def main():
    # Try to load the workbook
    try:
        wb = openpyxl.load_workbook(PRIMARY_PATH)
        target_path = PRIMARY_PATH
    except FileNotFoundError:
        print(f"File not found: {PRIMARY_PATH}")
        print("Creating new workbook...")
        wb = openpyxl.Workbook()
        target_path = PRIMARY_PATH

    # Get or create Sheet2 named "展示项"
    SHEET_NAME = "展示项"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.create_sheet(title=SHEET_NAME)
        print(f"Created new sheet: {SHEET_NAME}")

    build_sheet(ws)

    # Try to save to primary path, fallback if permission error
    try:
        wb.save(target_path)
        print(f"Saved successfully: {target_path}")
    except PermissionError:
        print(f"PermissionError saving to {target_path}")
        print(f"Saving to fallback: {FALLBACK_PATH}")
        wb.save(FALLBACK_PATH)
        print(f"Saved successfully: {FALLBACK_PATH}")


if __name__ == "__main__":
    main()
